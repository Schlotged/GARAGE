from tkinter import *
import tkinter
from tkinter import ttk
from docxtpl import DocxTemplate
import datetime
from tkinter import messagebox
from os import *
from docx2pdf import *
import tkinter as tk
from tkinter import  ttk
import openpyxl
from docxtpl import DocxTemplate
from docx import Document
from tkinter import filedialog


# Criar uma tela com tkinter
janela = tkinter.Tk()
janela.title("Gerador de orçamento")
janela.geometry("1000x650")
janela.configure(background="black")


note = ttk.Notebook(janela)
note.pack()

tab1=tkinter.Frame(note)
note.add(tab1, text="Orçamento")

# Limpar os campos da tela
def clear_item():
    quantidade_spinbox.delete(0, tkinter.END)
    quantidade_spinbox.insert(0, "1")
    descricao_entry.delete(0, tkinter.END)
    preco_unitario_spinbox.delete(0, tkinter.END)
    preco_unitario_spinbox.insert(0, "0.0")

# Adicionar os itens dos campos na tela
invoice_list = []
def add_item():
    qty = int(quantidade_spinbox.get()) 
    desc = descricao_entry.get()
    preco = float(preco_unitario_spinbox.get())
    line_total = qty*preco
    invoice_item = [qty, desc, preco, line_total]

    tree.insert("",0, values=invoice_item)
    clear_item()
    
    invoice_list.append(invoice_item)

# Limpar a tela para um novo documento
def new_invoice():
    first_name_entry.delete(0, tkinter.END)
    last_name_entry.delete(0, tkinter.END)
    telefone_entry.delete(0, tkinter.END)
    clear_item()
    treeview.delete(*treeview.get_children())
    
    invoice_list.clear()

# Função do botão para gerar o documento em word e converter em PDF
def gerar_documento():
    doc = DocxTemplate("Orcamento_direto_caixa_teste.docx")
    name = first_name_entry.get()+ "_" + last_name_entry.get()
    phone = telefone_entry.get()
    subtotal = sum(item[3] for item in invoice_list)
    salestax = 0.1
    total = subtotal*(1-salestax) 
    
    doc.render({"nome": name,
            "Telefone": phone,
            "invoice_list": invoice_list,
            "subtotal": subtotal,
            "taxavenda":str(salestax*100)+"%",
            "total": total})
    
    doc_name = "Novo_orçamento" + "_" + name + "_" + datetime.datetime.now().strftime("%Y-%m-%d-%H-%M") + ".docx"
    doc.save(doc_name)
    messagebox.showinfo("Documento gerado com sucesso", "Documento gerado com sucesso")
    new_invoice()
    convert(doc_name)
    messagebox.showinfo("Documento gerado com sucesso", "Documento PDF gerado com sucesso")

# Frame dos primeiros campos
posicao_cima_frame = ttk.Labelframe(tab1, text="Insira")
posicao_cima_frame.grid(row=0, column=0, padx=10, pady=10, sticky="ew")

# Campo do Primeiro nome
first_name_frame = tkinter.Label(posicao_cima_frame, text="Primeiro nome")
first_name_frame.grid(row= 0, column=0, padx=20, pady=10, sticky="nsew")
first_name_entry = tkinter.Entry(posicao_cima_frame)
first_name_entry.grid(row=1, column=0,padx=20, pady=10, sticky="nsew")

# Campo do Ultimo Nome
last_name_frame = tkinter.Label(posicao_cima_frame, text="Ultimo nome")
last_name_frame.grid(row=0, column=1, padx=20, pady=10, sticky="nsew")
last_name_entry = tkinter.Entry(posicao_cima_frame)
last_name_entry.grid(row=1, column=1,padx=20, pady=10, sticky="nsew")

# Campo do telefone na tela
telefone_label = tkinter.Label(posicao_cima_frame, text="Telefone")
telefone_label.grid(row=0, column=2,padx=20, pady=10, sticky="nsew")
telefone_entry = tkinter.Entry(posicao_cima_frame)
telefone_entry.grid(row=1, column=2,padx=20, pady=10, sticky="nsew")

# Campo da Quantidade na tela
quantidade_label = tkinter.Label(posicao_cima_frame, text="Quantidade")
quantidade_label.grid(row=2, column=0,padx=20, pady=10, sticky="nsew")
quantidade_spinbox = tkinter.Spinbox(posicao_cima_frame, from_=0, to=100)
quantidade_spinbox.grid(row=3, column=0, padx=20, pady=10, sticky="nsew")

# Campo da Descrição na tela
descricao_label = tkinter.Label(posicao_cima_frame, text="Descrição")
descricao_label.grid(row=2, column=1,padx=20, pady=10, sticky="nsew")
descricao_entry = tkinter.Entry(posicao_cima_frame)
descricao_entry.grid(row=3, column=1,padx=20, pady=10, sticky="nsew")

# Campo do Preço Unitario na tela
preco_unitario_label = tkinter.Label(posicao_cima_frame, text="Preco Unitario")
preco_unitario_label.grid(row=2, column=2,padx=20, pady=10, sticky="nsew")
preco_unitario_spinbox = tkinter.Spinbox(posicao_cima_frame, from_=0.0, to=1000, increment=0.1)
preco_unitario_spinbox.grid(row=3, column=2,padx=20, pady=10, sticky="nsew")

# Botão de Adicionar item
add_item_button = tkinter.Button(posicao_cima_frame, text="Adicionar Item", command=add_item)
add_item_button.grid(row=4, column=0, padx=20, pady=10, sticky="nsew")

gerador_frame = ttk.Labelframe(tab1, text="Tela de Dados")
gerador_frame.grid(row=4, column=0, padx=20, pady=10, sticky="news")

# Botão para gerar o documento word e consequentemente converter em PDF
save_invoice_button = ttk.Button(gerador_frame, text="Gerar Documento", command= gerar_documento)
save_invoice_button.grid(row=0, column=0, padx=20, pady=5)

# Botão parar limpar a tela e realizar um novo processo
new_invoice_button = ttk.Button(gerador_frame, text="Nova Fatura", command=new_invoice)
new_invoice_button.grid(row=0, column=1,padx=20, pady=5)

# Frame do Treeview na Tela
tree_frame = ttk.Labelframe(tab1, text="Tela de Dados")
tree_frame.grid(row=3, column=0, padx=20, pady=10)

# Parametrização da tela do treeview
columns = ("Quantidade", "Descricao", "Preco Unitario", "Total")
tree = ttk.Treeview(tree_frame, columns=columns, show="headings")
tree.grid(row=0, column=0, padx=50, pady=20)
tree.heading("Quantidade", text="Quantidade")
tree.heading("Descricao", text="Descricao")
tree.heading("Preco Unitario",text="Preco Unitario")
tree.heading("Total", text="Total")

# Segunda aba da tela do TKINTER
tab2 = tkinter.Frame(note)
note.add(tab2, text="Lançamento")

# Função para relacionar os dados com o treeview
def load_data():
    filepath = "base de dados_teste.xlsx"
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active
    
    list_values= list(sheet.values)

    for value_tuple in list_values[1:]: 
        treeview.insert('', tk.END, values=value_tuple)
    
    for col_name in columns:
        treeview.heading(col_name, text=col_name)

# Função do botão para insetir os dados dos campos na planilha e mostrar no treeview
selected_items = []
def insert_row():
    produto = produto_entry.get()
    valor_unidade = float(valor_entry.get()) 
    quantidade = int(qnt_spinbox.get())
    valor_total = valor_unidade*quantidade
    name = name_entry.get()
    age = int(age_spinbox.get())
    subscription_status = status_combobox.get()
    cliente = tipo_cliente_combobox.get()
    vendedor = name_vendedor_entry.get()
     
    path = "base de dados_teste.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    row_values = [name, age, subscription_status, cliente, produto, valor_unidade,quantidade, valor_total, vendedor]
    sheet.append(row_values)
    workbook.save(path)

    treeview.insert('', tk.END, values=row_values)


    name_entry.delete(0, "end")
    name_entry.insert(0, "Nome")
    age_spinbox.delete(0, "end")
    age_spinbox.insert(0, "Total dias fiado")
    status_combobox.set(combo_list[0])
    
     
    selected_item = {
        "Nome": name,
        "Dias": age,
        "Pagamento": subscription_status,
        "Cliente": tipo_cliente_combobox,
        "Produto": produto,
        "Valor unitario": valor_unidade,
        "Quantidade": quantidade,
        "Total": valor_total,
    }
    selected_items.append(selected_item)


def generate_word_document(): 
    filepath = "base de dados_teste.xlsx"
    workbook = openpyxl.load_workbook(filepath)
    name_dius = name_entry.get()
    template_path = "Orcamento_detalhado.docx"
    output_path = f"Caixa_" + name_dius + "_" + datetime.datetime.now().strftime('%Y-%m-%d-%H-%M') + ".docx"
    sheet = workbook.active
 
    
    list_values = list(sheet.values)

    doc = DocxTemplate(template_path)

    for value_tuple in list_values[1:]:
        name = value_tuple[0]
        age = value_tuple[1]
        subscription_status = value_tuple[2]
        cliente = value_tuple[3]
        produto = value_tuple[4]
        valor_unidade = value_tuple[5]
        quantidade = value_tuple[6]
        valor_total = value_tuple[7]
        vendedor = value_tuple[8]

        doc.render({
            "Nome": name,
            "Dias": age,
            "Vendedor": vendedor,
            "Pagamento": subscription_status,
            "Cliente": cliente,
            "Quantidade": quantidade,
            "Produto": produto,
            "valor_unitario": valor_unidade,
            "total": valor_total,
        })
        
    doc.save(output_path)
    messagebox.showinfo("Sucesso", "Documento gerado e salvo com sucesso!")


def janela_top_level():
    janela_dois = Toplevel()
    janela_dois.title("Relatório")
    janela_dois.configure(background="white")
    janela_dois.geometry("400x300")
    janela_dois.resizable(False, False)
    janela_dois.transient(janela)
    janela_dois.focus_force()
    janela_dois.grab_set()
     

combo_list = ["", "Pago a vista", "Crediario", "Parcelado cartão", "Dinheiro"]
combo_lista_dois = ["", "Fixo", "Novo", "Fiado"] 

Widgets_frame = ttk.Labelframe(tab2, text="Insira")
Widgets_frame.grid(row=0, column=0, padx=20, pady=5, sticky="w")

name_label = ttk.Label(Widgets_frame, text="Nome")
name_label.grid(row=0, column=0,padx=5, pady=(0, 5),sticky="ew")

name_entry = ttk.Entry(Widgets_frame)
name_entry.insert(0, "Nome")
name_entry.bind("<FocusIn>", lambda e: name_entry.delete("0", "end"))
name_entry.grid(row=1, column=0,padx=5, pady=(0, 5),sticky="ew")

produto_label = ttk.Label(Widgets_frame, text="Produto")
produto_label.grid(row=2, column=0,padx=5, pady=5,sticky="ew")

produto_entry = ttk.Entry(Widgets_frame)
produto_entry.insert(0, "Produto")
produto_entry.bind("<FocusIn>", lambda e: produto_entry.delete("0", "end"))
produto_entry.grid(row=3, column=0, padx=10, pady=5, sticky="ew")

tipo_cliente_label = ttk.Label(Widgets_frame, text="Tipo de Cliente")
tipo_cliente_label.grid(row=4, column=0, padx=5, pady=5, sticky="ew")

tipo_cliente_combobox = ttk.Combobox(Widgets_frame, text="Cliente", values=combo_lista_dois)
tipo_cliente_combobox.insert(0, "")
tipo_cliente_combobox.grid(row=5, column=0, padx=5, pady=5, sticky="ew")

valor_label = ttk.Label(Widgets_frame, text="Valor do Produto")
valor_label.grid(row=0, column=1, padx=5, pady=(0,5), sticky="ew")

valor_entry = ttk.Spinbox(Widgets_frame, from_=0.0, to=1000, increment=0.1)
valor_entry.grid(row=1, column=1, padx=5, pady=(0,5), sticky="ew")

tipo_pagamento_label = ttk.Label(Widgets_frame, text="Tipo de Pagamento")
tipo_pagamento_label.grid(row=2, column=1, padx=5, pady=5, sticky="ew")

status_combobox = ttk.Combobox(Widgets_frame, values=combo_list)
status_combobox.current(0)
status_combobox.grid(row=3, column=1, padx=10, pady=5, sticky="ew")

quantidade_label = ttk.Label(Widgets_frame, text="Quantidade")
quantidade_label.grid(row=0, column=2, padx=5, pady=(0, 5), sticky="nsew")

qnt_spinbox = ttk.Spinbox(Widgets_frame, from_=1, to=100)
qnt_spinbox.insert(0, "Quantidade")
qnt_spinbox.bind("<FocusIn>", qnt_spinbox.delete(0, "end"))
qnt_spinbox.grid(row=1, column=2, padx=5, pady=(0,5), sticky="nsew")

dias_label = ttk.Label(Widgets_frame, text="Prazo cobrança")
dias_label.grid(row=2, column=2, padx=10, pady=5, sticky="ew")

age_spinbox = ttk.Spinbox(Widgets_frame, from_=0, to=30)
age_spinbox.insert(0, "Total dias fiado")
age_spinbox.bind("<FocusIn>", lambda e: age_spinbox.delete("0", "end"))
age_spinbox.grid(row=3, column=2, padx=10, pady=5, sticky="ew")

name_vendedor_label = ttk.Label(Widgets_frame, text="Vendedor")
name_vendedor_label.grid(row=4, column=1, padx=5, pady=(0, 5), sticky="ew")

name_vendedor_entry = ttk.Entry(Widgets_frame)
name_vendedor_entry.insert(0, "Vendedor")
name_vendedor_entry.bind("<FocusIn>", lambda e: name_vendedor_entry.delete("0", "end"))
name_vendedor_entry.grid(row=5, column=1,padx=5, pady=(0, 5),sticky="ew") 
 
parte_de_baixo_frame = ttk.Labelframe(tab2, text="DADOS")
parte_de_baixo_frame.grid(row=4, column=0, padx=20, pady=10, sticky="w")

button = ttk.Button(parte_de_baixo_frame, text="Insira", command=insert_row)
button.grid(row=0, column=0,  padx=20, pady=5,sticky="nsew")

botao_gerar_documento = ttk.Button(parte_de_baixo_frame, text="Gerar")
botao_gerar_documento.grid(row=0, column=1,  padx=20, pady=5,sticky="nsew")


treeview_labelframe = ttk.LabelFrame(tab2, text="Tela")
treeview_labelframe.grid(row=3, column=0,columnspan=3, padx=20, pady=10)

treeFrame = ttk.Frame(treeview_labelframe)
treeFrame.grid(row=3, column=0, pady=10)
treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side="right", fill="y")


columns = ("Nome","Dias","Pagamento","Cliente","Produto", "Valor unitario", "Quantidade","Total", "Vendedor")
treeview = ttk.Treeview(treeFrame, show="headings", yscrollcommand=treeScroll.set, columns=columns, height=13)
treeview.column("Nome", width=100)
treeview.column("Dias", width=50)
treeview.column("Pagamento", width=100)
treeview.column("Cliente", width=100)
treeview.column("Produto", width=200)
treeview.column("Valor unitario", width=100)
treeview.column("Quantidade", width=100)
treeview.column("Total", width=100)
treeview.column("Vendedor", width=100)
treeview.pack()
treeScroll.config(command=treeview.yview)
load_data()


janela.mainloop()
